import streamlit as st
import io
from copy import copy
from concurrent.futures import ThreadPoolExecutor
from utils.sheets import (
    get_all_media,
    get_major_categories,
    get_sub_categories,
    get_creative_guides,
    build_media_cat_map,
    export_sheet_as_xlsx,
)
from utils.auth import get_current_user
from utils.ui import set_current_page
from st_click_detector import click_detector
import openpyxl
from openpyxl.utils import get_column_letter

set_current_page("creative_guide")
user = get_current_user()

st.markdown(
    "<div style='background:rgba(242,169,59,0.3); color:#7A4E0A; "
    "border:1px solid rgb(242,169,59); border-radius:8px; "
    "font-size:13px; font-weight:600; padding:9px 14px; margin-bottom:16px;'>"
    "<span style='display:inline-block; font-size:10.5px; font-weight:700; "
    "padding:2px 8px; border-radius:4px; margin-right:10px; letter-spacing:0.02em; "
    "vertical-align:middle; background:#F2A93B; color:#1C1200;'>공지</span>"
    "매체별 가이드 업데이트 진행 중 - 완료 시 별도 안내 예정"
    "</div>",
    unsafe_allow_html=True,
)

# ============================
# 데이터 로드
# ============================

guides = get_creative_guides()
all_media = get_all_media()
media_cat_map = build_media_cat_map()

guide_map = {}
for g in guides:
    guide_map.setdefault(g["media_name"], {})[g["product_name"]] = g

# ============================
# 필터 UI
# ============================

majors = get_major_categories()

col_f1, col_f2, col_or, col_f3 = st.columns([2, 2, 0.4, 3])

with col_f1:
    major_filter = st.selectbox(
        "대분류", [""] + majors, label_visibility="collapsed",
        key="cg_major", format_func=lambda x: "대분류 선택" if x == "" else x,
    )

with col_f2:
    if major_filter:
        subs = get_sub_categories(major_filter)
        sub_filter = st.selectbox(
            "중분류", [""] + subs, label_visibility="collapsed",
            key="cg_sub", format_func=lambda x: "중분류" if x == "" else x,
        )
    else:
        sub_filter = ""
        st.selectbox("중분류", ["중분류"], label_visibility="collapsed",
            key="cg_sub_empty", disabled=True)

with col_or:
    st.markdown(
        "<div style='text-align:center;color:var(--text-muted);font-size:12px;padding-top:8px;'>또는</div>",
        unsafe_allow_html=True,
    )

with col_f3:
    search_kw = st.text_input("검색", placeholder="🔍 매체명 직접 검색",
        label_visibility="collapsed", key="cg_search")

# ============================
# 필터 로직
# ============================

any_filter = bool(major_filter or search_kw)
all_media_names = sorted(set([m["name"] for m in all_media] + list(guide_map.keys())))


def passes_filter(name: str) -> bool:
    if search_kw and search_kw.lower() not in name.lower():
        return False
    if major_filter and media_cat_map.get(name, "") != major_filter:
        return False
    return True


filtered_names = [n for n in all_media_names if passes_filter(n)] if any_filter else []

if "cg_selected" not in st.session_state:
    st.session_state["cg_selected"] = {}

# ============================
# 매체/상품 목록
# ============================

if not any_filter:
    st.markdown(
        "<div style='margin:40px auto;max-width:480px;background:rgba(0,0,0,0.04);"
        "border-radius:12px;padding:24px 28px;opacity:0.6;text-align:center;'>"
        "<div style='font-size:14px;font-weight:600;margin-bottom:12px;'>Quick Guide</div>"
        "<div style='font-size:13px;color:var(--text-secondary);text-align:center;line-height:2;'>"
        "① 희망하는 매체를 카테고리에서 직접 선택하거나 검색<br>"
        "② 해당 상품 체크<br>"
        "③ 체크 완료된 파일 확인 후 다운로드 버튼 클릭!"
        "</div></div>",
        unsafe_allow_html=True,
    )

else:
    # 시안 C — Split Button (좌: 선택 / 우: 원본 이동)
    all_rows_html = []
    for media_name in filtered_names:
        products = guide_map.get(media_name, {})
        if not products:
            continue

        btn_parts = []
        for product_name, guide in sorted(products.items()):
            has_file = bool(guide.get("has_file"))
            sheet_url = guide.get("sheet_url", "")
            key = (media_name, product_name)
            is_on = st.session_state["cg_selected"].get(key, False)
            pid = f"{media_name}||{product_name}"

            if not has_file:
                btn_parts.append(
                    f"<span style='padding:6px 14px;font-size:13px;border-radius:8px;"
                    f"box-shadow:0 0 0 0.5px #666 inset;"
                    f"color:#999;opacity:0.45;cursor:not-allowed;"
                    f"display:inline-block;'>{product_name}</span>"
                )
            elif is_on:
                # 선택된 상태 (검은 배경) + 우측 링크 영역
                if sheet_url:
                    btn_parts.append(
                        f"<span style='display:inline-flex;align-items:stretch;"
                        f"background:#111;border-radius:8px;overflow:hidden;font-size:13px;line-height:1;'>"
                        f"<a href='#' id='{pid}' style='text-decoration:none;color:#fff;"
                        f"padding:7px 14px;cursor:pointer;display:inline-flex;align-items:center;'>"
                        f"✓ {product_name}</a>"
                        f"<span style='width:1px;background:rgba(255,255,255,0.25);'></span>"
                        f"<a href='{sheet_url}' target='_blank' rel='noopener' "
                        f"title='스프레드시트 열기' "
                        f"style='text-decoration:none;color:#fff;padding:7px 10px;"
                        f"display:inline-flex;align-items:center;'>↗</a>"
                        f"</span>"
                    )
                else:
                    btn_parts.append(
                        f"<a href='#' id='{pid}' style='text-decoration:none;'>"
                        f"<span style='padding:6px 14px;font-size:13px;border-radius:8px;"
                        f"background:#111;color:#fff;cursor:pointer;display:inline-block;'>"
                        f"✓ {product_name}</span></a>"
                    )
            else:
                # 미선택 상태 (밝은 배경) + 우측 링크 영역
                if sheet_url:
                    btn_parts.append(
                        f"<span style='display:inline-flex;align-items:stretch;"
                        f"box-shadow:0 0 0 0.5px #111 inset;border-radius:8px;overflow:hidden;"
                        f"font-size:13px;line-height:1;'>"
                        f"<a href='#' id='{pid}' style='text-decoration:none;color:#111;"
                        f"padding:7px 14px;cursor:pointer;display:inline-flex;align-items:center;'>"
                        f"{product_name}</a>"
                        f"<span style='width:1px;background:rgba(0,0,0,0.2);'></span>"
                        f"<a href='{sheet_url}' target='_blank' rel='noopener' "
                        f"title='스프레드시트 열기' "
                        f"style='text-decoration:none;color:#111;padding:7px 10px;"
                        f"display:inline-flex;align-items:center;'>↗</a>"
                        f"</span>"
                    )
                else:
                    btn_parts.append(
                        f"<a href='#' id='{pid}' style='text-decoration:none;'>"
                        f"<span style='padding:6px 14px;font-size:13px;border-radius:8px;"
                        f"box-shadow:0 0 0 0.5px #111 inset;color:#111;"
                        f"cursor:pointer;display:inline-block;'>{product_name}</span></a>"
                    )

        row_html = (
            "<div style='display:flex;align-items:center;gap:0;padding:10px 0;"
            "border-bottom:0.5px solid #e0e0e0;'>"
            f"<div style='flex:0 0 80px;font-size:13px;font-weight:600;'>{media_name}</div>"
            "<div style='width:1px;height:32px;background:#ccc;margin:0 12px;flex-shrink:0;'></div>"
            "<div style='display:flex;flex-wrap:wrap;gap:8px;flex:1;'>"
            + "".join(btn_parts) +
            "</div></div>"
        )
        all_rows_html.append(row_html)

    if all_rows_html:
        full_html = "<div>" + "".join(all_rows_html) + "</div>"
        clicked = click_detector(full_html, key="cg_det_all")

        last_key = "_cg_last_all"
        if clicked and "||" in clicked and clicked != st.session_state.get(last_key):
            st.session_state[last_key] = clicked
            mn, pn = clicked.split("||", 1)
            ck = (mn, pn)
            g = guide_map.get(mn, {}).get(pn)
            if g and bool(g.get("has_file")):
                st.session_state["cg_selected"][ck] = not st.session_state["cg_selected"].get(ck, False)
                st.rerun()

    # ============================
    # 선택 태그 + 다운로드 버튼
    # ============================

    selected = {k: v for k, v in st.session_state["cg_selected"].items() if v}

    if selected:
        tag_html = "".join(
            f"<span style='font-size:12px;padding:4px 12px;border-radius:20px;"
            f"box-shadow:0 0 0 0.5px #999 inset;"
            f"display:inline-block;margin:3px;'>"
            f"{mn} · {pn}</span>"
            for (mn, pn) in selected
        )
        st.markdown(
            f"<div style='margin-top:16px;margin-bottom:8px;'>{tag_html}</div>",
            unsafe_allow_html=True,
        )

    def _build_merged_workbook(sel_items):
        """선택 상품들 xlsx 병렬 다운로드 → openpyxl 병합."""
        # 1) 유효 URL만 추림
        fetch_tasks = []
        for (mn, pn) in sel_items:
            guide = guide_map.get(mn, {}).get(pn)
            if guide and guide.get("sheet_url"):
                fetch_tasks.append(((mn, pn), guide["sheet_url"]))

        # 2) 병렬 xlsx 다운로드 (매체별 네트워크 요청 동시)
        with ThreadPoolExecutor(max_workers=5) as ex:
            futures = {key: ex.submit(export_sheet_as_xlsx, url) for key, url in fetch_tasks}
            xlsx_results = [(key, f.result()) for key, f in futures.items()]

        # 3) 순차 병합 (CPU 작업)
        merged_wb = openpyxl.Workbook()
        merged_wb.remove(merged_wb.active)

        for (mn, pn), xlsx_bytes in xlsx_results:
            src_wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
            for sheet_name in src_wb.sheetnames:
                src_ws = src_wb[sheet_name]
                new_ws = merged_wb.create_sheet(title=sheet_name[:31])
                new_ws.sheet_view.showGridLines = False

                for row in src_ws.iter_rows():
                    for cell in row:
                        new_ws[cell.coordinate].value = cell.value
                        if cell.has_style:
                            new_ws[cell.coordinate].font = copy(cell.font)
                            new_ws[cell.coordinate].fill = copy(cell.fill)
                            new_ws[cell.coordinate].border = copy(cell.border)
                            new_ws[cell.coordinate].alignment = copy(cell.alignment)
                            new_ws[cell.coordinate].number_format = cell.number_format

                for col in range(1, src_ws.max_column + 1):
                    letter = get_column_letter(col)
                    src_dim = src_ws.column_dimensions.get(letter)
                    if src_dim and src_dim.width:
                        new_ws.column_dimensions[letter].width = src_dim.width
                    else:
                        max_len = 0
                        for r in src_ws.iter_rows(min_col=col, max_col=col,
                                                   max_row=min(src_ws.max_row, 100)):
                            for c in r:
                                if c.value:
                                    max_len = max(max_len, len(str(c.value)))
                        new_ws.column_dimensions[letter].width = max(
                            8.5, min(max_len * 1.2, 80)
                        )

                for row_dim in src_ws.row_dimensions.values():
                    if row_dim.height:
                        new_ws.row_dimensions[row_dim.index].height = row_dim.height

                for merge in src_ws.merged_cells.ranges:
                    new_ws.merge_cells(str(merge))

        buf = io.BytesIO()
        merged_wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # 선택 조건 변경 시 이전 준비된 파일 폐기
    sel_key = tuple(sorted(selected.keys()))
    if st.session_state.get("_cg_download_key") != sel_key:
        st.session_state.pop("_cg_download_bytes", None)
        st.session_state["_cg_download_key"] = sel_key

    @st.fragment
    def download_fragment():
        if "_cg_download_bytes" in st.session_state:
            # 준비 완료 → 실제 다운로드 버튼
            st.download_button(
                "📥 제작가이드 통합 다운로드",
                data=st.session_state["_cg_download_bytes"],
                file_name="통합_제작가이드.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        else:
            # 준비 전 → 준비 요청 버튼
            if st.button(
                "📦 제작가이드 다운로드 준비",
                type="primary",
                use_container_width=True,
                key="cg_prepare_btn",
            ):
                with st.spinner("파일 다운로드 준비 중, 잠시만 기다려주세요..."):
                    try:
                        data = _build_merged_workbook(list(selected.keys()))
                        st.session_state["_cg_download_bytes"] = data
                        st.rerun()
                    except Exception as e:
                        st.error(f"다운로드 준비 중 오류: {e}")

    download_fragment()
