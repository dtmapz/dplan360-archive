import io
import re
import html
from collections import Counter
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.auth import get_current_user
from utils.ui import set_current_page
from utils.sheets import get_budget_history, get_budget_adv

set_current_page("budget_reference")
user = get_current_user()

# 캠페인 식별 기준 (내부 로직 전용 — 광고주/브랜드/캠페인명은 화면·엑셀에 노출하지 않음)
GROUP_COLS = ["대업종", "소업종", "광고주", "브랜드", "대행사", "캠페인명", "대행사 발행월"]
# 화면/엑셀에 실제로 노출되는 컬럼 (익명 라벨 "구분"이 광고주/브랜드/캠페인명을 대체)
DISPLAY_COLS = ["구분", "매체사", "상품", "광고수주액", "비중", "집행월"]
EMPTY_PRODUCT = "—"


def _norm_month(s: str) -> str:
    """대행사 발행월 표기 정규화 → 'YYYY.MM'. 파싱 불가면 원문 유지."""
    s = (s or "").strip()
    if not s:
        return s
    m = re.match(r"^(\d{4})[.\-/](\d{1,2})(?:\.0+)?$", s)
    if m:
        return f"{m.group(1)}.{int(m.group(2)):02d}"
    # gspread가 이미 float로 캐스팅한 흔적 (예: '2025.1' → 원본 '2025.10'일 가능성)
    # 소수부 1자리이고 두 자릿수로 복원할 수 없으므로 그대로 유지
    return s


def _campaign_label(rank: int) -> str:
    """0-indexed rank → 익명 캠페인 라벨 (A, B, ... Z, AA, AB, ...)."""
    return get_column_letter(rank + 1)


# ============================================================
# 데이터 로드 + 조인
# ============================================================
@st.cache_data(ttl=600, show_spinner=False)
def _load_joined() -> pd.DataFrame:
    hist = pd.DataFrame(get_budget_history())
    adv = pd.DataFrame(get_budget_adv())

    if hist.empty:
        return hist

    for col in ("광고주", "브랜드"):
        if col in hist.columns:
            hist[col] = hist[col].astype(str).str.strip()
    for col in ("광고주명", "브랜드명"):
        if col in adv.columns:
            adv[col] = adv[col].astype(str).str.strip()

    if not adv.empty:
        adv_slim = adv[["광고주명", "브랜드명", "대업종", "소업종"]].drop_duplicates(
            subset=["광고주명", "브랜드명"], keep="first"
        )
        hist = hist.merge(
            adv_slim,
            how="left",
            left_on=["광고주", "브랜드"],
            right_on=["광고주명", "브랜드명"],
        ).drop(columns=["광고주명", "브랜드명"], errors="ignore")

    if "대행사 발행월" in hist.columns:
        hist["대행사 발행월"] = hist["대행사 발행월"].astype(str).str.strip().map(_norm_month)

    # 상품 컬럼 정규화 (헤더가 "상품" / "상품명" / "상품(정리)" 어느 쪽이든 대응)
    for alt in ("상품(정리)", "상품명"):
        if "상품" not in hist.columns and alt in hist.columns:
            hist["상품"] = hist[alt]
            break
    if "상품" in hist.columns:
        hist["상품"] = hist["상품"].astype(str).str.strip()
        hist["상품"] = hist["상품"].where(hist["상품"] != "", EMPTY_PRODUCT)
    else:
        hist["상품"] = EMPTY_PRODUCT

    if "광고수주액" in hist.columns:
        hist["광고수주액"] = pd.to_numeric(
            hist["광고수주액"].astype(str).str.replace(",", "").str.strip(),
            errors="coerce",
        ).fillna(0).astype(int)

    return hist


try:
    df_all = _load_joined()
except Exception as e:
    st.error(f"데이터 로드 실패: {e}")
    st.stop()

if df_all.empty:
    st.info("budget_history 시트에 데이터가 없습니다.")
    st.stop()


# ============================================================
# 필터 UI
# ============================================================
majors = sorted(
    [x for x in df_all.get("대업종", pd.Series(dtype=str)).dropna().unique() if str(x).strip()]
)

col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns([1.6, 2.2, 1.1, 1.1, 0.9])

with col_f1:
    sel_major = st.selectbox("대업종 (필수)", ["-- 선택 --"] + majors, index=0)

sub_options: list[str] = []
if sel_major and sel_major != "-- 선택 --":
    sub_options = sorted(
        [
            x for x in df_all[df_all["대업종"] == sel_major]
            .get("소업종", pd.Series(dtype=str))
            .dropna()
            .unique()
            if str(x).strip()
        ]
    )

with col_f2:
    sel_subs = st.multiselect("소업종 (선택)", sub_options, default=[])

# 연/월 옵션 — 선택한 업종 범위 내에서만 노출
_scope = df_all.copy()
if sel_major and sel_major != "-- 선택 --":
    _scope = _scope[_scope["대업종"] == sel_major]
if sel_subs:
    _scope = _scope[_scope["소업종"].isin(sel_subs)]

_month_series = _scope.get("대행사 발행월", pd.Series(dtype=str)).dropna().astype(str)
year_options = sorted({m.split(".")[0] for m in _month_series if re.match(r"^\d{4}\.\d{2}$", m)})

with col_f3:
    sel_years = st.multiselect("연도 (선택)", year_options, default=[])

if sel_years:
    _month_series = _month_series[_month_series.str[:4].isin(sel_years)]
month_options = sorted({m.split(".")[1] for m in _month_series if re.match(r"^\d{4}\.\d{2}$", m)})

with col_f4:
    sel_months = st.multiselect("월 (선택)", month_options, default=[])

with col_f5:
    # 라벨 높이만큼 spacer 추가해 버튼을 드롭다운 하단선에 정렬
    st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
    search = st.button("조회", type="primary", use_container_width=True)

if search:
    if not sel_major or sel_major == "-- 선택 --":
        st.warning("대업종을 선택해주세요.")
        st.stop()

    df = df_all.copy()
    df = df[df["대업종"] == sel_major]
    if sel_subs:
        df = df[df["소업종"].isin(sel_subs)]
    if sel_years:
        df = df[df["대행사 발행월"].astype(str).str[:4].isin(sel_years)]
    if sel_months:
        df = df[df["대행사 발행월"].astype(str).str[-2:].isin(sel_months)]

    st.session_state["_budget_result"] = df.reset_index(drop=True)


# ============================================================
# 캠페인(그룹) 단위 집계
# ============================================================
def _group_iter(df: pd.DataFrame):
    """(group_key, group_df, total_amount) 캠페인 총액 desc.
    하나의 group = 하나의 '캠페인' (광고주/브랜드/대행사/캠페인명/집행월 기준 식별, 화면엔 미노출)."""
    grouped = df.groupby(GROUP_COLS, dropna=False, sort=False)
    items = []
    for key, gdf in grouped:
        total = int(gdf["광고수주액"].sum())
        items.append((key, gdf.reset_index(drop=True), total))
    items.sort(key=lambda x: x[2], reverse=True)
    return items


def _compute_media_summary(groups) -> list[dict]:
    """매체사별 활용 캠페인 수 + 비중(매체 활용 건수 합계 대비)."""
    counter = Counter()
    for _key, gdf, _total in groups:
        for media in gdf["매체사"].dropna().astype(str).str.strip().unique():
            if media:
                counter[media] += 1
    total_uses = sum(counter.values())
    rows = []
    for media, cnt in counter.most_common():
        pct = (cnt / total_uses * 100) if total_uses else 0
        rows.append({"매체사": media, "활용 캠페인 수": cnt, "비중": pct})
    return rows, total_uses


def _compute_combo_insights(groups) -> dict:
    """매체 조합 / 매체·상품 조합 / 최다 매체 활용 캠페인 인사이트."""
    combo_counter = Counter()      # frozenset(매체사...) → 캠페인 수
    media_product_counter = Counter()  # (매체사, 상품) → 캠페인 수
    most_diverse = None            # (media_count, media_list)

    for _key, gdf, _total in groups:
        media_list = [m for m in gdf["매체사"].dropna().astype(str).str.strip().unique() if m]
        if not media_list:
            continue
        combo_counter[frozenset(media_list)] += 1

        mp_pairs = set()
        for _, row in gdf.iterrows():
            media = str(row.get("매체사") or "").strip()
            product = str(row.get("상품") or "").strip() or EMPTY_PRODUCT
            if media:
                mp_pairs.add((media, product))
        for pair in mp_pairs:
            media_product_counter[pair] += 1

        if most_diverse is None or len(media_list) > most_diverse[0]:
            most_diverse = (len(media_list), sorted(media_list))

    top_combo = combo_counter.most_common(1)
    top_mp = media_product_counter.most_common(1)

    return {
        "top_combo": (sorted(top_combo[0][0]), top_combo[0][1]) if top_combo else None,
        "top_media_product": top_mp[0] if top_mp else None,
        "most_diverse": most_diverse,
    }


# ============================================================
# 결과 렌더링
# ============================================================
def _fmt_won(v: int) -> str:
    return f"{int(v):,}"


def _fmt_pct(v: float) -> str:
    return f"{v:.1f}%"


def _render_summary_html(rows: list[dict], total_uses: int, major: str, subs: list[str]) -> str:
    css = """
    <style>
    .bref-sum { border-collapse: collapse; width: 100%; font-size: 12px; background: #fff; }
    .bref-sum th, .bref-sum td {
        border: 1px solid #E5E5E5; padding: 8px 10px; text-align: center; vertical-align: middle;
    }
    .bref-sum thead th { background: #0B0B0B; color: #fff; font-weight: 600; padding: 10px; }
    .bref-sum td.label { text-align: center; font-weight: 600; }
    .bref-sum td.merged { background: #fafafa; }
    .bref-sum td.num { text-align: right; font-variant-numeric: tabular-nums; }
    .bref-sum tr.total td { background: #FFF8E1; font-weight: 700; border-top: 1.5px solid #F2A93B; }
    </style>
    """
    show_sub = bool(subs)
    sub_label = ", ".join(subs)
    n = len(rows)

    headers = ["대업종"]
    if show_sub:
        headers.append("소업종")
    headers += ["매체사", "비중", "활용 캠페인 수"]
    head = "<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"

    def _row(i: int, r: dict) -> str:
        tds = []
        if i == 0:
            tds.append(f'<td class="label merged" rowspan="{n}">{html.escape(major)}</td>')
            if show_sub:
                tds.append(f'<td class="label merged" rowspan="{n}">{html.escape(sub_label)}</td>')
        tds.append(f'<td class="label">{html.escape(r["매체사"])}</td>')
        tds.append(f'<td class="num">{_fmt_pct(r["비중"])}</td>')
        tds.append(f'<td class="num">{r["활용 캠페인 수"]}</td>')
        return "<tr>" + "".join(tds) + "</tr>"

    body = "".join(_row(i, r) for i, r in enumerate(rows))
    lead_colspan = 3 if show_sub else 2  # 대업종(+소업종) + 매체사
    total_row = (
        f'<tr class="total"><td class="label" colspan="{lead_colspan}">Total</td><td class="num">100%</td>'
        f'<td class="num">{total_uses}</td></tr>'
    )
    return css + f'<table class="bref-sum"><thead>{head}</thead><tbody>{body}{total_row}</tbody></table>'


def _render_combo_html(insights: dict) -> str:
    def chips(items: list[str]) -> str:
        return '<span style="color:#999; margin:0 4px;">+</span>'.join(
            f'<span style="background:#0B0B0B; color:#fff; font-size:12px; font-weight:600; '
            f'padding:4px 10px; border-radius:6px;">{html.escape(m)}</span>'
            for m in items
        )

    cards = []

    if insights["top_combo"]:
        media_list, cnt = insights["top_combo"]
        cards.append(("가장 많이 활용된 매체 조합", chips(media_list), f"{cnt}건"))
    if insights["top_media_product"]:
        (media, product), cnt = insights["top_media_product"]
        pair_html = (
            f'<span style="background:#0B0B0B; color:#fff; font-size:12px; font-weight:600; '
            f'padding:4px 10px; border-radius:6px;">{html.escape(media)}</span>'
            f'<span style="background:#f0f0f0; color:#111; font-size:12px; font-weight:600; '
            f'padding:4px 10px; border-radius:6px; margin-left:6px; border:1px solid #ddd;">{html.escape(product)}</span>'
        )
        cards.append(("가장 많이 활용된 매체·상품", pair_html, f"{cnt}건"))
    if insights["most_diverse"]:
        n, media_list = insights["most_diverse"]
        cards.append(("가장 다양한 매체를 활용한 조합", chips(media_list), f"{n}개 매체"))

    if not cards:
        return ""

    card_html = "".join(
        f'<div style="flex:1; min-width:220px; background:#fff; border:1px solid #E5E5E5; '
        f'border-radius:12px; padding:16px; display:flex; flex-direction:column; gap:12px;">'
        f'<div style="font-size:12px; font-weight:600; color:#666; min-height:34px;">{title}</div>'
        f'<div style="display:flex; flex-wrap:wrap; align-items:center; gap:4px;">{chip_html}</div>'
        f'<div style="margin-top:auto; font-size:20px; font-weight:700; color:#7A4E0A;">{count}</div>'
        f'</div>'
        for title, chip_html, count in cards
    )
    return f'<div style="display:flex; gap:14px; flex-wrap:wrap;">{card_html}</div>'


def _render_preview_html(groups) -> str:
    css = """
    <style>
    .bref-tbl { border-collapse: collapse; width: 100%; font-size: 12px; background: #fff; }
    .bref-tbl th, .bref-tbl td {
        border: 1px solid #E5E5E5; padding: 8px 10px; text-align: center; vertical-align: middle;
    }
    .bref-tbl thead th {
        background: #0B0B0B; color: #fff; font-weight: 600; padding: 10px;
    }
    .bref-tbl td.num { text-align: right; font-variant-numeric: tabular-nums; }
    .bref-tbl td.merged { background: #fafafa; }
    .bref-tbl td.anon { color: #999; font-style: italic; }
    .bref-tbl tr.total td {
        background: #FFF8E1; font-weight: 700; border-top: 1.5px solid #F2A93B;
    }
    </style>
    """
    head = "<tr>" + "".join(f"<th>{c}</th>" for c in DISPLAY_COLS) + "</tr>"
    body_rows = []

    for rank, (key, gdf, total) in enumerate(groups):
        key_map = dict(zip(GROUP_COLS, key))
        n = len(gdf)
        label = _campaign_label(rank)

        for i, row in gdf.iterrows():
            tds = []
            if i == 0:
                tds.append(f'<td class="merged anon" rowspan="{n}">캠페인 {label}</td>')
            tds.append(f'<td>{html.escape(str(row.get("매체사") or ""))}</td>')
            product = str(row.get("상품") or "").strip() or EMPTY_PRODUCT
            tds.append(f'<td>{html.escape(product)}</td>')
            amt = int(row.get("광고수주액") or 0)
            tds.append(f'<td class="num">{_fmt_won(amt)}</td>')
            pct = (amt / total * 100) if total else 0
            tds.append(f'<td class="num">{_fmt_pct(pct)}</td>')
            if i == 0:
                tds.append(f'<td class="merged" rowspan="{n}">{html.escape(str(key_map.get("대행사 발행월") or ""))}</td>')
            body_rows.append("<tr>" + "".join(tds) + "</tr>")

        body_rows.append(
            f'<tr class="total">'
            f'<td colspan="4">Total : 캠페인 {label}</td>'
            f'<td class="num">{_fmt_won(total)}</td>'
            f'<td></td>'
            f'</tr>'
        )

    return css + f'<table class="bref-tbl"><thead>{head}</thead><tbody>{"".join(body_rows)}</tbody></table>'


def _build_excel(groups) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "campaigns"

    header_fill = PatternFill("solid", fgColor="0B0B0B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill("solid", fgColor="FFF8E1")
    total_font = Font(bold=True)
    merged_fill = PatternFill("solid", fgColor="FAFAFA")
    thin = Side(border_style="thin", color="E5E5E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # 구분(1) | 매체사(2) | 상품(3) | 광고수주액(4) | 비중(5) | 집행월(6)
    for col_idx, name in enumerate(DISPLAY_COLS, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    r = 2
    for rank, (key, gdf, total) in enumerate(groups):
        key_map = dict(zip(GROUP_COLS, key))
        n = len(gdf)
        start_row = r
        label = f"캠페인 {_campaign_label(rank)}"

        for i, row in gdf.iterrows():
            if i == 0:
                cell = ws.cell(row=r, column=1, value=label)
                cell.fill = merged_fill
                cell = ws.cell(row=r, column=6, value=key_map.get("대행사 발행월") or "")
                cell.fill = merged_fill

            ws.cell(row=r, column=2, value=str(row.get("매체사") or ""))
            product = str(row.get("상품") or "").strip() or EMPTY_PRODUCT
            ws.cell(row=r, column=3, value=product)
            amt = int(row.get("광고수주액") or 0)
            ac = ws.cell(row=r, column=4, value=amt)
            ac.number_format = "#,##0"
            ac.alignment = right
            pct = (amt / total) if total else 0
            pc = ws.cell(row=r, column=5, value=pct)
            pc.number_format = "0.0%"
            pc.alignment = right
            r += 1

        # Total 행
        total_row_idx = r
        tc = ws.cell(row=total_row_idx, column=1, value=f"Total : {label}")
        for col_idx in range(1, 7):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
        ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=3)
        tc.alignment = center
        ac = ws.cell(row=total_row_idx, column=4, value=total)
        ac.number_format = "#,##0"
        ac.alignment = right
        ac.fill = total_fill
        ac.font = total_font
        ws.merge_cells(start_row=total_row_idx, start_column=5, end_row=total_row_idx, end_column=6)
        r += 1

        # 구분(1)·집행월(6) rowspan 병합 (2건 이상일 때)
        end_row = start_row + n - 1
        if end_row > start_row:
            for col_idx in [1, 6]:
                ws.merge_cells(
                    start_row=start_row, start_column=col_idx,
                    end_row=end_row, end_column=col_idx,
                )
                ws.cell(row=start_row, column=col_idx).alignment = center

    for row_cells in ws.iter_rows(min_row=1, max_row=r - 1, min_col=1, max_col=6):
        for c in row_cells:
            c.border = border
            if c.alignment is None or c.alignment.horizontal is None:
                c.alignment = center

    widths = [10, 18, 18, 16, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


df_result = st.session_state.get("_budget_result")
if df_result is not None:
    groups = _group_iter(df_result)
    n_campaigns = len(groups)

    st.markdown(
        f"<div style='margin-top:14px; padding:12px 16px; background:#f0f0f0; "
        f"border:0.5px solid #E5E5E5; border-radius:8px; font-size:13px;'>"
        f"조회 조건으로 <b>{n_campaigns:,}개 캠페인</b>이 확인되었습니다."
        f"</div>",
        unsafe_allow_html=True,
    )

    if n_campaigns == 0:
        st.stop()

    # ---- 업종 미디어믹스 요약 ----
    scope_label = " · ".join([sel_major] + sel_subs) if sel_subs else sel_major
    summary_rows, total_uses = _compute_media_summary(groups)
    st.markdown(
        "<div style='margin-top:22px; font-size:13px; font-weight:600; margin-bottom:6px;'>"
        "업종 미디어믹스 요약 <span style='font-size:11px; font-weight:400; color:#666; margin-left:6px;'>"
        f"선택 범위 내 전체 캠페인 집계 · {html.escape(scope_label)} · 매체 활용 건수 합계 {total_uses}건 기준"
        "</span></div>",
        unsafe_allow_html=True,
    )
    st.markdown(_render_summary_html(summary_rows, total_uses, sel_major, sel_subs), unsafe_allow_html=True)

    # ---- 매체 활용 조합 ----
    combo_insights = _compute_combo_insights(groups)
    combo_html = _render_combo_html(combo_insights)
    if combo_html:
        st.markdown(
            "<div style='margin-top:22px; font-size:13px; font-weight:600; margin-bottom:6px;'>"
            "매체 활용 조합 <span style='font-size:11px; font-weight:400; color:#666; margin-left:6px;'>"
            "선택 범위 내 캠페인 단위 조합 분석</span></div>",
            unsafe_allow_html=True,
        )
        st.markdown(combo_html, unsafe_allow_html=True)

    # ---- 상위 3개 캠페인 ----
    st.markdown(
        "<div style='margin-top:22px; font-size:13px; font-weight:600; margin-bottom:6px;'>"
        "미리보기 <span style='font-size:11px; font-weight:400; color:#666; margin-left:6px;'>"
        "상위 3개 캠페인 · 총액 순 · 광고주/브랜드/캠페인명 비공개</span></div>",
        unsafe_allow_html=True,
    )
    st.markdown(_render_preview_html(groups[:3]), unsafe_allow_html=True)

    excel_bytes = _build_excel(groups)
    fname_parts = [sel_major]
    if sel_subs:
        fname_parts.append("-".join(sel_subs)[:30])
    fname = f"budget_history_{'_'.join(fname_parts)}.xlsx"

    st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)
    st.download_button(
        label=f"엑셀 다운로드 (전체 {n_campaigns:,}개 캠페인)",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
